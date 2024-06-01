import csv
import openpyxl
import os
import sys

from openpyxl import load_workbook

def excel_to_csv(excel_file):
    # Separa la ruta del archivo y su extensión
    file_path, file_extension = os.path.splitext(excel_file)

    # Reemplaza la extensión por ".csv"
    csv_file = file_path + ".csv"

    workbook = load_workbook(filename=excel_file)
    sheet = workbook.active
    csv_data = []

    # Lee datos desde el Excel
    for value in sheet.iter_rows(values_only=True):
        csv_data.append(list(value))

    # Escribe en CSV
    with open(csv_file, 'w', newline='', encoding='utf-8') as csv_file_obj:
        writer = csv.writer(csv_file_obj, delimiter=',')
        for line in csv_data:
            writer.writerow(line)

if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("Uso: python excel_to_csv.py <archivo_excel>")
        sys.exit(1)

    excel_file = sys.argv[1]
    excel_to_csv(excel_file)
